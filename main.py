# main.py  (aiogram >= 3.7,<3.9)
import os
import io
import json
import asyncio
import logging
import datetime
import calendar
from pathlib import Path
from collections import defaultdict
from typing import Dict, Any, Iterable

from aiogram import Bot, Dispatcher, Router, F
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BotCommand,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    BufferedInputFile,
    BotCommandScopeDefault,
    BotCommandScopeChat,
)
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ================== КОНФИГ ==================
RAW_TOKEN = os.getenv("BOT_TOKEN", "")
BOT_TOKEN = (
    RAW_TOKEN.replace("\u00A0", " ").replace("\r", "").replace("\n", "").strip().strip('"').strip("'")
)
if not BOT_TOKEN or ":" not in BOT_TOKEN:
    raise RuntimeError(f"BOT_TOKEN выглядит неверно. RAW={RAW_TOKEN!r}")

OWNER_ID  = 104653853
ADMIN_IDS = [104653853, 1155243378]  # можно расширять

# ===== Папка для постоянного хранилища =====
DATA_DIR = Path("/data")
DATA_DIR.mkdir(parents=True, exist_ok=True)

# файлы данных (будут храниться в /data на Railway)
EMP_FILE   = DATA_DIR / "employees.json"
SHIFT_FILE = DATA_DIR / "shifts.json"

# МСК
MSK = ZoneInfo("Europe/Moscow")

# ===== Нормы для ОТЧЁТА (Excel) — НЕ МЕНЯЕМ =====
START_NORM = datetime.time(8, 0)
START_OK_TILL = datetime.time(8, 10)
END_NORM = datetime.time(17, 30)
END_OK_TILL = datetime.time(17, 40)

# ===== Допуски ТОЛЬКО для вопросов в боте =====
PROMPT_EARLY_OK_FROM = datetime.time(7, 45)  # раньше — спросим причину
PROMPT_START_OK_TILL = datetime.time(8, 10)  # позже — спросим причину
PROMPT_END_OK_TILL   = datetime.time(17, 45) # позже — спросим причину

# ================== ИНИЦИАЛИЗАЦИЯ ==================
logging.basicConfig(level=logging.INFO)
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())
router = Router()
dp.include_router(router)

# ================== ПАМЯТЬ/ДАННЫЕ ==================
# shifts_by_date["YYYY-MM-DD"][user_id] = {...}
shifts_by_date: Dict[str, Dict[int, Dict[str, Any]]] = defaultdict(dict)

# EMPLOYEES: id -> {"name": str, "active": bool}
DEFAULT_EMPLOYEES = {
    str(OWNER_ID): {"name": "OWNER", "active": True},
}
EMPLOYEES: Dict[int, Dict[str, Any]] = {}

# ожидаем причину (по пользователю) — "start_early"|"start_late"|"end_early"|"end_late"
pending_reason: Dict[int, str] = {}

# ================== УТИЛИТЫ ==================
def msk_now() -> datetime.datetime:
    return datetime.datetime.now(MSK)

def today_key() -> str:
    return msk_now().date().isoformat()

def fmt_hm(dt: datetime.datetime | None) -> str:
    if not dt: return "—"
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=MSK)
    return dt.astimezone(MSK).strftime("%H:%M")

def is_weekend(date: datetime.date) -> bool:
    return calendar.weekday(date.year, date.month, date.day) >= 5

def fio(uid: int) -> str:
    meta = EMPLOYEES.get(uid)
    return (meta or {}).get("name") if meta else f"Неизвестный ({uid})"

def is_admin(uid: int) -> bool:
    return uid in ADMIN_IDS or uid == OWNER_ID

def is_allowed(uid: int) -> bool:
    if uid == OWNER_ID or uid in ADMIN_IDS:
        return True
    meta = EMPLOYEES.get(uid)
    return bool(meta and meta.get("active", True))

def ensure_allowed(message: Message) -> bool:
    uid = message.from_user.id
    if not is_allowed(uid):
        asyncio.create_task(message.answer("Нет доступа. Обратитесь к администратору."))
        return False
    return True

def today_shift(uid: int) -> Dict[str, Any]:
    return shifts_by_date[today_key()].setdefault(uid, {})

# ---- I/O сотрудников и смен
def load_employees() -> dict[int, Dict[str, Any]]:
    if EMP_FILE.exists():
        try:
            raw = json.loads(EMP_FILE.read_text("utf-8"))
            result: dict[int, Dict[str, Any]] = {}
            for k, v in raw.items():
                uid = int(k)
                if isinstance(v, str):  # обратная совместимость
                    result[uid] = {"name": v, "active": True}
                else:
                    name = str(v.get("name", f"ID {uid}"))
                    active = bool(v.get("active", True))
                    result[uid] = {"name": name, "active": active}
            return result
        except Exception as e:
            logging.exception("Не удалось прочитать employees.json: %s", e)
    # создаём по умолчанию
    EMP_FILE.write_text(json.dumps(DEFAULT_EMPLOYEES, ensure_ascii=False, indent=2), "utf-8")
    # привести ключи к int
    return {int(k): v for k, v in DEFAULT_EMPLOYEES.items()}

def save_employees() -> None:
    out = {str(k): {"name": v.get("name",""), "active": bool(v.get("active", True))} for k, v in EMPLOYEES.items()}
    EMP_FILE.write_text(json.dumps(out, ensure_ascii=False, indent=2), "utf-8")

def dt_to_iso(dt: datetime.datetime | None) -> str | None:
    return dt.astimezone(MSK).isoformat() if dt else None

def dt_from_iso(s: str | None) -> datetime.datetime | None:
    if not s: return None
    return datetime.datetime.fromisoformat(s)

def save_shifts() -> None:
    data_out: dict[str, dict[str, dict[str, Any]]] = {}
    for day, users in shifts_by_date.items():
        data_out[day] = {}
        for uid, d in users.items():
            data_out[day][str(uid)] = {
                "start": dt_to_iso(d.get("start")),
                "end": dt_to_iso(d.get("end")),
                "start_reason": d.get("start_reason"),
                "end_reason": d.get("end_reason"),
                "comment": d.get("comment"),
                "comment_done": d.get("comment_done"),
            }
    SHIFT_FILE.write_text(json.dumps(data_out, ensure_ascii=False, indent=2), "utf-8")

def load_shifts() -> None:
    if not SHIFT_FILE.exists(): return
    try:
        data_in = json.loads(SHIFT_FILE.read_text("utf-8"))
    except Exception as e:
        logging.exception("Не удалось прочитать shifts.json: %s", e)
        return
    for day, users in data_in.items():
        shifts_by_date[day] = {}
        for uid_str, d in users.items():
            uid = int(uid_str)
            shifts_by_date[day][uid] = {
                "start": dt_from_iso(d.get("start")),
                "end": dt_from_iso(d.get("end")),
                "start_reason": d.get("start_reason"),
                "end_reason": d.get("end_reason"),
                "comment": d.get("comment"),
                "comment_done": d.get("comment_done"),
            }

# загрузка при старте
EMPLOYEES = load_employees()
load_shifts()

# ================== КЛАВИАТУРЫ ==================
user_buttons = [
    [KeyboardButton(text="Смену начал 🏭"), KeyboardButton(text="Смену закончил 🏡")],
    [KeyboardButton(text="Мой статус📍"), KeyboardButton(text="Инструкция 📖")],
]
admin_buttons = user_buttons + [[KeyboardButton(text="Отчет 📈"), KeyboardButton(text="Статус смены 🛠")]]

def kb(uid: int) -> ReplyKeyboardMarkup:
    base = admin_buttons if is_admin(uid) else user_buttons
    if uid == OWNER_ID:  # только Owner видит меню сотрудников
        base = [*base, [KeyboardButton(text="Сотрудники ⚙️")]]
    return ReplyKeyboardMarkup(keyboard=base, resize_keyboard=True)

owner_menu_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="❇️ Добавить сотрудника"), KeyboardButton(text="🔴 Деактивировать")],
        [KeyboardButton(text="🟢 Активировать"), KeyboardButton(text="🗑 Удалить сотрудника")],
        [KeyboardButton(text="📋 Список сотрудников")],
        [KeyboardButton(text="⬅️ Назад")],
    ],
    resize_keyboard=True
)

# ================== КОМАНДЫ ==================
@router.message(CommandStart())
async def cmd_start(message: Message):
    if not ensure_allowed(message): return
    await message.answer("Добро пожаловать в бот учёта рабочего времени DUSBERG ⏱️", reply_markup=kb(message.from_user.id))

@router.message(Command("whoami"))
async def cmd_whoami(message: Message):
    if not ensure_allowed(message): return
    uid = message.from_user.id
    role = "OWNER" if uid == OWNER_ID else ("ADMIN" if is_admin(uid) else "USER")
    meta = EMPLOYEES.get(uid, {})
    active_str = "активен" if meta.get("active", True) else "неактивен"
    await message.answer(
        f"Ты: <b>{role}</b> ({active_str})\n"
        f"ФИО: <b>{fio(uid)}</b>\n"
        f"ID: <code>{uid}</code>",
        reply_markup=kb(uid)
    )

# ================== OWNER-МЕНЮ «Сотрудники ⚙️» ==================
class EmpStates(StatesGroup):
    wait_add = State()
    wait_del = State()
    wait_deactivate = State()
    wait_activate = State()

@router.message(F.text == "Сотрудники ⚙️")
async def owner_menu(message: Message):
    if message.from_user.id != OWNER_ID: return
    await message.answer(
        "Меню сотрудников:\n"
        "• «❇️ Добавить сотрудника» — пришлите: <code>123456789 Иванов И.И.</code>\n"
        "• «🔴 Деактивировать» — пришлите: <code>123456789</code>\n"
        "• «🟢 Активировать» — пришлите: <code>123456789</code>\n"
        "• «🗑 Удалить сотрудника» — пришлите: <code>123456789</code>\n"
        "• «📋 Список сотрудников» — показать текущий справочник.",
        reply_markup=owner_menu_kb
    )

@router.message(F.text == "⬅️ Назад")
async def owner_back(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    await state.clear()
    await message.answer("Главное меню.", reply_markup=kb(message.from_user.id))

@router.message(F.text == "📋 Список сотрудников")
async def owner_list(message: Message):
    if message.from_user.id != OWNER_ID: return
    if not EMPLOYEES:
        await message.answer("Справочник пуст.", reply_markup=owner_menu_kb); return

    # сортировка по ФИО (name), затем по id для стабильности
    items = sorted(EMPLOYEES.items(), key=lambda kv: (kv[1].get("name","").lower(), kv[0]))

    chunk = []
    for uid, meta in items:
        status = "🟢 активен" if meta.get("active", True) else "🔴 неактивен"
        chunk.append(f"{uid}: {meta.get('name','')} — {status}")
        if len(chunk) == 50:
            await message.answer("\n".join(chunk), reply_markup=owner_menu_kb)
            chunk = []
    if chunk:
        await message.answer("\n".join(chunk), reply_markup=owner_menu_kb)

@router.message(F.text == "❇️ Добавить сотрудника")
async def owner_add_start(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    await state.set_state(EmpStates.wait_add)
    await message.answer("Пришлите строку: <code>123456789 Иванов И.И.</code>", reply_markup=owner_menu_kb)

@router.message(EmpStates.wait_add, F.text)
async def owner_add_do(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    text = (message.text or "").strip()
    parts = text.split(maxsplit=1)
    if len(parts) < 2:
        return await message.answer("Нужно и ID, и ФИО. Пример: <code>123456789 Иванов И.И.</code>", reply_markup=owner_menu_kb)
    try:
        new_id = int(parts[0])
    except ValueError:
        return await message.answer("ID должен быть числом.", reply_markup=owner_menu_kb)
    name = parts[1].strip().strip('"').strip("'")
    if not name:
        return await message.answer("Пустое имя.", reply_markup=owner_menu_kb)
    EMPLOYEES[new_id] = {"name": name, "active": True}
    save_employees()
    await state.clear()
    await message.answer(f"Добавлен: {new_id} — {name} (🟢 активен)", reply_markup=owner_menu_kb)

@router.message(F.text == "🗑 Удалить сотрудника")
async def owner_del_start(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    await state.set_state(EmpStates.wait_del)
    await message.answer("Пришлите ID сотрудника. Пример: <code>123456789</code>", reply_markup=owner_menu_kb)

@router.message(EmpStates.wait_del, F.text)
async def owner_del_do(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    text = (message.text or "").strip()
    try:
        uid_del = int(text)
    except ValueError:
        return await message.answer("ID должен быть числом.", reply_markup=owner_menu_kb)
    if uid_del == OWNER_ID:
        return await message.answer("Нельзя удалить OWNER.", reply_markup=owner_menu_kb)
    if EMPLOYEES.pop(uid_del, None) is None:
        await state.clear()
        return await message.answer("Такого ID нет в справочнике.", reply_markup=owner_menu_kb)
    save_employees()
    await state.clear()
    await message.answer(f"Удалён: {uid_del}", reply_markup=owner_menu_kb)

@router.message(F.text == "🔴 Деактивировать")
async def owner_deactivate_start(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    await state.set_state(EmpStates.wait_deactivate)
    await message.answer("Пришлите ID сотрудника для деактивации. Пример: <code>123456789</code>", reply_markup=owner_menu_kb)

@router.message(EmpStates.wait_deactivate, F.text)
async def owner_deactivate_do(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    text = (message.text or "").strip()
    try:
        uid_deact = int(text)
    except ValueError:
        return await message.answer("ID должен быть числом.", reply_markup=owner_menu_kb)
    if uid_deact == OWNER_ID:
        return await message.answer("Нельзя деактивировать OWNER.", reply_markup=owner_menu_kb)
    meta = EMPLOYEES.get(uid_deact)
    if not meta:
        await state.clear()
        return await message.answer("Такого ID нет в справочнике.", reply_markup=owner_menu_kb)
    meta["active"] = False
    save_employees()
    await state.clear()
    await message.answer(f"Деактивирован: {uid_deact} — {meta.get('name','')} (🔴 неактивен)", reply_markup=owner_menu_kb)

@router.message(F.text == "🟢 Активировать")
async def owner_activate_start(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    await state.set_state(EmpStates.wait_activate)
    await message.answer("Пришлите ID сотрудника для активации. Пример: <code>123456789</code>", reply_markup=owner_menu_kb)

@router.message(EmpStates.wait_activate, F.text)
async def owner_activate_do(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID: return
    text = (message.text or "").strip()
    try:
        uid_act = int(text)
    except ValueError:
        return await message.answer("ID должен быть числом.", reply_markup=owner_menu_kb)
    meta = EMPLOYEES.get(uid_act)
    if not meta:
        await state.clear()
        return await message.answer("Такого ID нет в справочнике.", reply_markup=owner_menu_kb)
    meta["active"] = True
    save_employees()
    await state.clear()
    await message.answer(f"Активирован: {uid_act} — {meta.get('name','')} (🟢 активен)", reply_markup=owner_menu_kb)

# ================== БИЗНЕС-ЛОГИКА СМЕН ==================
@router.message(F.text == "Смену начал 🏭")
async def handle_start(message: Message):
    if not ensure_allowed(message): return
    uid = message.from_user.id
    now = msk_now()
    shift = today_shift(uid)

    if shift.get("start") and shift.get("end") is None:
        await message.answer("Смена уже начата. Сначала заверши текущую.", reply_markup=kb(uid))
        return

    shift["start"] = now
    shift["end"] = None
    shift["start_reason"] = None
    shift["end_reason"] = None
    shift["comment"] = None
    pending_reason.pop(uid, None)
    save_shifts()

    t = now.time()
    if is_weekend(now.date()):
        pending_reason[uid] = "start_early"
        await message.answer("Сегодня выходной. Укажи причину начала смены.", reply_markup=kb(uid))
    elif t < PROMPT_EARLY_OK_FROM:
        pending_reason[uid] = "start_early"
        await message.answer("Смена начата слишком рано. Укажи причину.", reply_markup=kb(uid))
    elif t > PROMPT_START_OK_TILL:
        pending_reason[uid] = "start_late"
        await message.answer("Смена начата позже. Укажи причину.", reply_markup=kb(uid))
    else:
        await message.answer("Смена начата. Продуктивного дня!", reply_markup=kb(uid))

@router.message(F.text == "Смену закончил 🏡")
async def handle_end(message: Message):
    if not ensure_allowed(message): return
    uid = message.from_user.id
    now = msk_now()
    shift = today_shift(uid)

    if not shift.get("start"):
        await message.answer("Смена ещё не начата.", reply_markup=kb(uid))
        return
    if shift.get("end"):
        await message.answer("Смена уже завершена.", reply_markup=kb(uid))
        return

    shift["end"] = now
    pending_reason.pop(uid, None)
    save_shifts()

    t = now.time()
    if t < END_NORM:
        pending_reason[uid] = "end_early"
        await message.answer("Смена завершена слишком рано. Укажи причину.", reply_markup=kb(uid))
    elif t > PROMPT_END_OK_TILL:
        pending_reason[uid] = "end_late"
        await message.answer("Смена завершена позже. Укажи причину переработки.", reply_markup=kb(uid))
    else:
        await message.answer("Спасибо! Хорошего отдыха!", reply_markup=kb(uid))

# ================== ИНФО-КНОПКИ (с учётом эмодзи) ==================
@router.message(F.text.in_({"Мой статус", "Мой статус📍"}))
async def handle_status(message: Message):
    if not ensure_allowed(message): return
    uid = message.from_user.id
    data = shifts_by_date.get(today_key(), {}).get(uid)
    if not data:
        await message.answer("Смена не начата.", reply_markup=kb(uid))
        return

    lines = [
        f"Смена начата в: {fmt_hm(data.get('start'))}",
        f"Смена завершена в: {fmt_hm(data.get('end'))}",
    ]
    if data.get("start_reason"):
        lines.append(f"Причина отклонения (начало): {data['start_reason']}")
    if data.get("end_reason"):
        lines.append(f"Причина отклонения (завершение): {data['end_reason']}")
    if data.get("comment"):
        lines.append(f"Комментарий: {data['comment']}")
    await message.answer("\n".join(lines), reply_markup=kb(uid))

@router.message(F.text.in_({"Инструкция", "Инструкция 📖"}))
async def handle_help(message: Message):
    if not ensure_allowed(message): return
    await message.answer(
        "Для регистрации времени начала смены нажми - Смену начал 🏭 \n\n"
        "Для регистрации времени завершения смены нажми - Смену закончил 🏡\n\n"
        "Если бот спрашивает причину — ответь одним текстовым сообщением - это сохранится как причина отклонения.\n\n"
        "Дополнительные пояснения по смене можно прислать отдельным сообщением — это общий комментарий.",
        reply_markup=kb(message.from_user.id)
    )

@router.message(F.text.in_({"Статус смены", "Статус смены 🛠"}))
async def handle_shift_status(message: Message):
    if not ensure_allowed(message): return
    if not is_admin(message.from_user.id):
        await message.answer("Нет доступа.", reply_markup=kb(message.from_user.id))
        return

    day = today_key()
    day_data = shifts_by_date.get(day, {})
    if not day_data:
        await message.answer("Сегодня смен нет.", reply_markup=kb(message.from_user.id))
        return

    # список uid сегодняшних отметившихся, отсортированный по ФИО
    sorted_uids = sorted(day_data.keys(), key=lambda u: fio(u).lower())

    lines = []
    for uid in sorted_uids:
        data = day_data[uid]
        s = fmt_hm(data.get("start"))
        e = fmt_hm(data.get("end"))
        who = fio(uid)

        row = [f"{who}: начата в {s}, завершена в {e}"]

        reasons = []
        if data.get("start_reason"):
            reasons.append(f"начало — {data['start_reason']}")
        if data.get("end_reason"):
            reasons.append(f"завершение — {data['end_reason']}")

        if reasons:
            row.append("⚠️ Причина отклонения: " + "; ".join(reasons))

        lines.append("\n".join(row))

    await message.answer("\n\n".join(lines), reply_markup=kb(message.from_user.id))

# ================== ОТЧЁТ ПО ДИАПАЗОНУ (XLSX) ==================
class ReportStates(StatesGroup):
    waiting_period = State()

def daterange_inclusive(d1: datetime.date, d2: datetime.date) -> Iterable[datetime.date]:
    step = 1 if d1 <= d2 else -1
    cur = d1
    while True:
        yield cur
        if cur == d2: break
        cur = cur + datetime.timedelta(days=step)

def parse_date(s: str) -> datetime.date | None:
    s = (s or "").strip()
    # ISO: YYYY-MM-DD
    try:
        y, m, d = s.split("-")
        return datetime.date(int(y), int(m), int(d))
    except Exception:
        pass
    # RU: DD.MM.YY или DD.MM.YYYY
    try:
        d, m, y = s.split(".")
        if len(y) == 2:
            y = 2000 + int(y)  # 25 -> 2025
        else:
            y = int(y)
        return datetime.date(int(y), int(m), int(d))
    except Exception:
        return None

def calc_minutes(a: datetime.time, b: datetime.time) -> int:
    dt_a = datetime.datetime.combine(datetime.date.today(), a)
    dt_b = datetime.datetime.combine(datetime.date.today(), b)
    return int((dt_b - dt_a).total_seconds() // 60)

def deviation_columns(start_dt: datetime.datetime | None, end_dt: datetime.datetime | None) -> tuple[int,int,int,int]:
    """(раньше_начало, позже_начало, раньше_конец, позже_конец) в минутах (>=0) — для ОТЧЁТА"""
    early_start = late_start = early_end = late_end = 0
    if start_dt:
        st_local = start_dt.astimezone(MSK).time()
        if st_local < START_NORM:
            early_start = calc_minutes(st_local, START_NORM)
        if st_local > START_OK_TILL:
            late_start = calc_minutes(START_OK_TILL, st_local)
    if end_dt:
        en_local = end_dt.astimezone(MSK).time()
        if en_local < END_NORM:
            early_end = calc_minutes(en_local, END_NORM)
        if en_local > END_OK_TILL:
            late_end = calc_minutes(END_OK_TILL, en_local)
    return early_start, late_start, early_end, late_end

def minutes_between(start_dt: datetime.datetime | None, end_dt: datetime.datetime | None) -> int:
    if not start_dt or not end_dt: return 0
    a = start_dt.astimezone(MSK); b = end_dt.astimezone(MSK)
    if b < a: return 0
    return int((b - a).total_seconds() // 60)

def build_xlsx_bytes(date_from: datetime.date, date_to: datetime.date) -> bytes:
    wb = Workbook()
    ws_shifts = wb.active; ws_shifts.title = "Смены"
    ws_daily = wb.create_sheet("Свод по дням")
    ws_emps  = wb.create_sheet("Сотрудники")
    ws_params= wb.create_sheet("Параметры")

    # Заголовки
    ws_shifts.append([
        "Дата","Сотрудник","ID","Начало","Конец",
        "Раннее начало, мин","Позднее начало, мин","Раннее завершение, мин","Позднее завершение, мин",
        "Длительность, мин","Длительность, ч","Выходной",
        "Причина отклонения начала смены","Причина отклонения завершения смены","Комментарий"
    ])
    ws_daily.append([
        "Дата","Сотрудник","ID","Начало","Конец",
        "Раннее начало, мин","Позднее начало, мин","Раннее завершение, мин","Позднее завершение, мин",
        "Длительность, мин","Длительность, ч","Выходной"
    ])

    # Сотрудники — по алфавиту
    ws_emps.append(["ID","Сотрудник","Статус"])
    for uid, meta in sorted(EMPLOYEES.items(), key=lambda kv: (kv[1].get("name","").lower(), kv[0])):
        ws_emps.append([uid, meta.get("name",""), "активен" if meta.get("active", True) else "неактивен"])

    ws_params.append(["Параметр","Значение"])
    ws_params.append(["Часовой пояс","Europe/Moscow"])
    ws_params.append(["Норма начала","08:00"])
    ws_params.append(["Допустимо до (начало)","08:10"])
    ws_params.append(["Норма конца","17:30"])
    ws_params.append(["Допустимо до (конец)","17:40"])
    ws_params.append(["Период отчёта", f"{date_from.isoformat()} — {date_to.isoformat()}"])
    ws_params.append(["В отчёт включены все сотрудники, в том числе неактивные.", "Да"])

    # Данные: по каждому дню включаем ВСЕХ сотрудников (сортировка по имени)
    for day in daterange_inclusive(date_from, date_to):
        key = day.isoformat()
        day_data = shifts_by_date.get(key, {})
        weekend = "Да" if is_weekend(day) else "Нет"

        for uid, meta in sorted(EMPLOYEES.items(), key=lambda kv: (kv[1].get("name","").lower(), kv[0])):
            name = meta.get("name","")
            data = day_data.get(uid, None)

            if data:
                start_dt = data.get("start"); end_dt = data.get("end")
                start_str = fmt_hm(start_dt); end_str = fmt_hm(end_dt)
                early_start, late_start, early_end, late_end = deviation_columns(start_dt, end_dt)
                work_min = minutes_between(start_dt, end_dt)
                work_hours = round(work_min/60, 2)
                start_reason = data.get("start_reason") or ""
                end_reason   = data.get("end_reason") or ""
                comment      = data.get("comment") or ""
            else:
                start_str = end_str = "—"
                early_start = late_start = early_end = late_end = 0
                work_min = 0; work_hours = 0
                start_reason = end_reason = comment = ""

            ws_shifts.append([
                day, name, uid, start_str, end_str,
                early_start, late_start, early_end, late_end,
                work_min, work_hours, weekend,
                start_reason, end_reason, comment
            ])

            ws_daily.append([
                day, name, uid, start_str, end_str,
                early_start, late_start, early_end, late_end,
                work_min, work_hours, weekend
            ])

    # Оформление
    def fit_columns(ws):
        widths = {}
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row, 1):
                s = "" if cell is None else str(cell)
                widths[i] = max(widths.get(i, 0), len(s))
        for i, w in widths.items():
            ws.column_dimensions[get_column_letter(i)].width = min(max(10, w + 2), 40)

    for ws in (ws_shifts, ws_daily, ws_emps, ws_params):
        fit_columns(ws)
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ======== FSM отчёта ========
class ReportStates(StatesGroup):
    waiting_period = State()

@router.message(F.text.in_({"Отчет", "Отчет 📈"}))
async def ask_report_period(message: Message, state: FSMContext):
    if not ensure_allowed(message): return
    if not is_admin(message.from_user.id):
        await message.answer("Нет доступа.", reply_markup=kb(message.from_user.id))
        return
    await state.set_state(ReportStates.waiting_period)
    await message.answer(
        "Введите период дат (включительно):\n"
        "• Один день: <code>20.08.2025</code> или <code>20.08.25</code>\n"
        "• Диапазон: <code>01.08.2025 20.08.2025</code> (или ISO: <code>2025-08-01 2025-08-20</code>)\n"
        "Для отмены: /cancel"
    )

@router.message(Command("cancel"))
async def cancel_report(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Отменено.", reply_markup=kb(message.from_user.id))

@router.message(ReportStates.waiting_period, F.text)
async def handle_report_period(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("Нет доступа.", reply_markup=kb(message.from_user.id))
        await state.clear(); return

    parts = (message.text or "").strip().split()
    if len(parts) == 1:
        d1 = parse_date(parts[0]); d2 = d1
    elif len(parts) == 2:
        d1 = parse_date(parts[0]); d2 = parse_date(parts[1])
    else:
        d1 = d2 = None

    if not d1 or not d2:
        await message.answer(
            "Неверный формат. Примеры:\n"
            "• <code>20.08.2025</code> или <code>20.08.25</code>\n"
            "• <code>01.08.2025 20.08.2025</code> или <code>2025-08-01 2025-08-20</code>"
        )
        return

    if d2 < d1:
        d1, d2 = d2, d1

    if (d2 - d1).days > 92:
        await message.answer("Слишком длинный период (>92 дней). Сократите интервал.")
        await state.clear(); return

    try:
        xlsx = build_xlsx_bytes(d1, d2)
        fname = f"Отчёт_{d1.isoformat()}_{d2.isoformat()}.xlsx" if d1 != d2 else f"Отчёт_{d1.isoformat()}.xlsx"
        await message.answer_document(
            BufferedInputFile(xlsx, filename=fname),
            caption=f"Отчёт за период {d1.isoformat()} — {d2.isoformat()} (МСК).",
            reply_markup=kb(message.from_user.id)
        )
    except Exception as e:
        logging.exception("Ошибка формирования отчёта: %s", e)
        await message.answer("Не удалось сформировать отчёт. Проверьте данные и попробуйте ещё раз.")
    finally:
        await state.clear()

# ================== DEBUG ==================
from aiogram.filters import Command

@router.message(Command("debug_touch"))
async def debug_touch(message: Message):
    # только OWNER может дергать отладку
    if message.from_user.id != OWNER_ID:
        return await message.answer("Нет доступа.")
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        now = datetime.datetime.now().isoformat(timespec="seconds")
        EMP_FILE.write_text(f"debug employees at {now}\n", encoding="utf-8")
        SHIFT_FILE.write_text(f"debug shifts at {now}\n", encoding="utf-8")
        await message.answer("ok: wrote to /data")
    except Exception as ex:
        await message.answer(f"write error: {ex!r}")

@router.message(Command("debug_files"))
async def debug_files(message: Message):
    # только OWNER может дергать отладку
    if message.from_user.id != OWNER_ID:
        return await message.answer("Нет доступа.")
    try:
        e, s = EMP_FILE, SHIFT_FILE
        txt = (
            f"/data exists: {DATA_DIR.exists()}\n"
            f"{e.name}: exists={e.exists()} size={(e.stat().st_size if e.exists() else 0)} path={e}\n"
            f"{s.name}: exists={s.exists()} size={(s.stat().st_size if s.exists() else 0)} path={s}\n"
        )
        await message.answer(txt)
    except Exception as ex:
        await message.answer(f"error: {ex!r}")


# ================== СВОБОДНЫЙ ТЕКСТ (причины/комментарии) ==================
@router.message(~Command())  # важно: не перехватываем команды со слешем
async def handle_comment_or_reason(message: Message):
    if not ensure_allowed(message): return
    uid = message.from_user.id
    txt = (message.text or "").strip()
    if not txt:
        return

    # если ждём причину — сохраняем и добавляем хвостовую фразу
    reason_flag = pending_reason.get(uid)
    if reason_flag:
        shift = shifts_by_date.get(today_key(), {}).get(uid)
        if not shift:
            pending_reason.pop(uid, None)
            return

        if reason_flag in ("start_early", "start_late"):
            shift["start_reason"] = txt
            tail = " Продуктивного дня!"
        elif reason_flag in ("end_early", "end_late"):
            shift["end_reason"] = txt
            tail = " Хорошего отдыха!"
        else:
            tail = ""

        pending_reason.pop(uid, None)
        save_shifts()
        await message.answer("Спасибо! Причина зафиксирована." + tail, reply_markup=kb(uid))
        return

    # иначе — это общий комментарий к смене
    shift = shifts_by_date.get(today_key(), {}).get(uid)
    if not shift:
        return
    if shift.get("start") and not shift.get("end") and not shift.get("comment"):
        shift["comment"] = txt
        save_shifts()
        await message.answer("Комментарий сохранён. Продуктивного дня!", reply_markup=kb(uid))
    elif shift.get("end") and not shift.get("comment_done"):
        shift["comment_done"] = True
        save_shifts()
        await message.answer("Комментарий сохранен. Хорошего отдыха!", reply_markup=kb(uid))


# ================== ЗАПУСК ==================
async def main():
    try:
        me = await bot.get_me()
        logging.info("Авторизован как @%s (id=%s)", me.username, me.id)

        base_cmds = [
            BotCommand(command="start", description="Запуск бота"),
            BotCommand(command="whoami", description="Показать мою роль"),
            BotCommand(command="cancel", description="Отменить ввод периода"),
            # debug-команды можно не публиковать в меню
        ]
        await bot.set_my_commands(base_cmds, scope=BotCommandScopeDefault())
        await bot.set_my_commands(base_cmds, scope=BotCommandScopeChat(chat_id=OWNER_ID))

        await dp.start_polling(bot)
    except Exception as e:
        logging.exception("Старт не удался: %s", e)
    finally:
        try:
            save_shifts()
        finally:
            await bot.session.close()


if __name__ == "__main__":
    asyncio.run(main())
